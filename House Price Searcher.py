#!/usr/bin/env python
# coding: utf-8

# In[4]:


import re
data = "My postcode at home is EX20 1EH, my term postcode is YO10 5DD and YO10 3DU, EX22 6DJ, EX22 6NH "

postcode_finder = re.compile(r'(?:[A-Za-z]\d ?\d[A-Za-z]{2})|(?:[A-Za-z][A-Za-z\d]\d ?\d[A-Za-z]{2})|(?:[A-Za-z]{2}\d{2} ?\d[A-Za-z]{2})|(?:[A-Za-z]\d[A-Za-z] ?\d[A-Za-z]{2})|(?:[A-Za-z]{2}\d[A-Za-z] ?\d[A-Za-z]{2})')
postcode_searcher = postcode_finder.findall(data)
for i in postcode_searcher:
    print(i)

    


# In[6]:


import webbrowser, requests, pyautogui, time
from selenium import webdriver
browser = webdriver.Firefox()
browser.get("https://www.rightmove.co.uk/property-for-sale/find.html?locationIdentifier=POSTCODE%5E303577&radius=5.0&propertyTypes=&includeSSTC=false&mustHave=&dontShow=&furnishTypes=&keywords=")
pyautogui.click(265, 168)
pyautogui.press('backspace', presses = 8)
search = browser.find_element_by_class_name('input')


for i in postcode_searcher:
    try:
        print(i)
        search = browser.find_element_by_class_name('input')
        search.click
        search.send_keys(i)
        search.send_keys(u'\ue007')
        time.sleep(2)
        pyautogui.press('down', presses = 50, interval = 0.1)
        pyautogui.click(179, 118)
        pyautogui.hotkey('command', 'a')
        pyautogui.press('backspace')
    
    except pyautogui.FailSafeException:
        global postcode
        postcode = (i)
        print("The Postcode is: " + postcode)
        break

        
browser.get('https://online.mystery-shoppers.co.uk/document.asp?alias=MSLogin')   
username = browser.find_element_by_id('username')
username.send_keys('daniellynch654@yahoo.co.uk')

password = browser.find_element_by_id('password')
password.send_keys('Slithroat98')
password.send_keys(u'\ue007')
time.sleep(5)
pyautogui.click(165, 190)
pyautogui.click(205, 265)
time.sleep(3)

Id = browser.find_element_by_id('Login')
Id = Id.get_attribute('value')
Address_1 = browser.find_element_by_id('address1')
Address_1 = Address_1.get_attribute('value')
Address_2 = browser.find_element_by_id('address2')
Address_2 = Address_2.get_attribute('value')
Town = browser.find_element_by_id('city')
Town = Town.get_attribute('value')
Address = (Address_1  + " " + Address_2 + " " + " " + Town + " " +  postcode)

browser.get('https://www.zoopla.co.uk/')
time.sleep(1)
pyautogui.click(1125, 735)
house_price_check = browser.find_element_by_id('search-tabs-house-prices')
house_price_check.click()
zoopla_search = browser.find_element_by_id('search-input-location')
zoopla_search.send_keys(Address)
zoopla_search.send_keys(u'\ue007')





# In[7]:


import openpyxl, os
os.chdir("/Users/daniellynch/Downloads/")
wb = openpyxl.load_workbook("Example work sheet.xlsx")
ws = wb['Sheet1']
Address = (Address_1  + '\n' + Address_2 + "\n"  + Town + "\n" +  postcode)
print(Address)

Dummy_name_list = []
Email_list = []
Real_name_list = []
i = 0

for row in ws['B']:
    Dummy_name_list.append(row.value)
    
for row in ws['C']:
    Email_list.append(row.value)

for row in ws['E']:
    Real_name_list.append(row.value)
    

for cell in ws ['D']:
    i +=1
    Location = cell.value
    
    if Location == 'Okehampton':
        print("Match found " + "D" + str(i))
        Location_cell_number = ("F" + str(i))
        id_cell_number = ("G" + str(i))
        ws[Location_cell_number] = Address
        ws[id_cell_number] = Id
        print("Cells appended ")
        break
        


ws2 = wb['Sheet2']
x = 0

for row in ws2.iter_rows(min_row=1, max_col= x, max_row=1):
    for cell in row:
        x+=1
        if cell.value == "Okehampton":
            print("Match found " + "Row 1: Column: " + str(x))
            ws2.cell(row =2, column=x).value = Dummy_name_list[i-1]
            ws2.cell(row =3, column=x).value = Email_list[i-1]
            ws2.cell(row=4, column=x).value = Real_name_list[i-1]
            ws2.cell(row=5, column=x).value = Address
            ws2.cell(row=6, column=x).value = Id
            break
            
ws3 = wb['Sheet3']
Walk_in_details = []
Booked_visit_details = []
for cell, cell2 in zip(ws3['A'], ws3 ['B']):
    Booked_visit_details.append(cell.value)
    Walk_in_details.append(cell2.value)

y = 1

for row in ws2.iter_rows(min_row=7, min_col= 2, max_row=7, max_col = 7):
    for cell in row:
        y+=1
        if cell.value == 'Booked appointment':
            ws2.cell(row = 8, column = y).value = Booked_visit_details[1]
            ws2.cell(row = 9, column = y).value = Booked_visit_details[2]
            ws2.cell(row = 10, column = y).value = Booked_visit_details[3]
        elif cell.value == 'Walk in':
            ws2.cell(row = 8, column = y).value = Walk_in_details[1]
            ws2.cell(row = 9, column = y).value = Walk_in_details[2]
            ws2.cell(row = 10, column = y).value = Walk_in_details[3]
        else:
            pass
        
            
                         
                
    
    
            


wb.save("Example work sheet.xlsx")
print("Finished")


# In[ ]:


