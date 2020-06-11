#!/usr/bin/env python
# coding: utf-8

# In[10]:


import openpyxl, os
os.chdir("/Users/daniellynch/Downloads/")
wb = openpyxl.load_workbook("examplespreadsheet.xlsx")
ws = wb.active
column_A = []
column_B = []
for cell in ws["A"]:
    column_A.append(cell.value)
print (column_A)
for cell in ws["B"]:
    column_B.append(cell.value)
print(column_B)
wb = openpyxl.Workbook()
ws = wb.active
ws.append(column_A)
ws.append(column_B)

wb.save("invertedspreadsheet.xlsx")


# In[ ]:




